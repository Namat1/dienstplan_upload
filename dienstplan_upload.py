import streamlit as st
import pandas as pd
from zipfile import ZipFile
from datetime import datetime, date, time as datetime_time
import tempfile
import os
import time
import ftplib
import threading
import re
import socket
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from ftplib import FTP, FTP_TLS, error_perm
from dotenv import load_dotenv
from urllib.parse import quote, urlsplit, urlunsplit, parse_qsl, urlencode
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


# Manche Touren-Exceldateien enthalten beschädigte oder ungewöhnliche
# Kopf-/Fußzeilen. openpyxl ignoriert diese ohnehin; die Warnung würde bei
# vielen Dateien nur das Streamlit-Protokoll überfluten.
warnings.filterwarnings(
    "ignore",
    message=r"Cannot parse header or footer so it will be ignored",
    category=UserWarning,
)

# .env laden
load_dotenv()
FTP_HOST = os.getenv("FTP_HOST")
FTP_USER = os.getenv("FTP_USER")
FTP_PASS = os.getenv("FTP_PASS")
FTP_BASE_DIR = os.getenv("FTP_BASE_DIR", "/")
FTP_USE_TLS = os.getenv("FTP_TLS", "0") == "1"
FTP_PARALLEL = int(os.getenv("FTP_PARALLEL", "3"))
FTP_TIMEOUT = max(5, min(int(os.getenv("FTP_TIMEOUT", "15")), 30))
FTP_DOWNLOAD_MAX_SECONDS = max(15, int(os.getenv("FTP_DOWNLOAD_MAX_SECONDS", "45")))
DIENSTPLAN_CSV_URL = os.getenv("DIENSTPLAN_CSV_URL", "").strip()
HTTP_DOWNLOAD_TIMEOUT = max(10, min(int(os.getenv("HTTP_DOWNLOAD_TIMEOUT", "30")), 120))

# Deutsche Wochentage
wochentage_deutsch_map = {
    "Monday": "Montag",
    "Tuesday": "Dienstag",
    "Wednesday": "Mittwoch",
    "Thursday": "Donnerstag",
    "Friday": "Freitag",
    "Saturday": "Samstag",
    "Sunday": "Sonntag"
}

def get_plan_kw(start_sonntag):
    """Kalenderwoche der Planwoche Sonntag bis Samstag.

    Maßgeblich ist der Montag innerhalb dieser Planwoche. Dadurch wird
    der Jahreswechsel korrekt behandelt: 28.12.2025 bis 03.01.2026
    gehört zur Kalenderwoche 01/2026 und nicht zu Kalenderwoche 53.
    """
    montag = pd.Timestamp(start_sonntag) + pd.Timedelta(days=1)
    return int(montag.isocalendar().week)

def _new_ftp():
    """Eine frische, eingeloggte FTP(S)-Verbindung im Passive-Mode."""
    ftp = FTP_TLS() if FTP_USE_TLS else FTP()
    ftp.connect(FTP_HOST, 21, timeout=FTP_TIMEOUT)
    ftp.login(FTP_USER, FTP_PASS)
    if FTP_USE_TLS:
        ftp.prot_p()
    ftp.set_pasv(True)
    return ftp

def ensure_ftp_dirs(ftp, remote_dir, created_cache):
    remote_dir = remote_dir.replace("\\", "/")
    if remote_dir in created_cache:
        return
    path_built = ""
    for part in remote_dir.split("/"):
        if not part:
            continue
        path_built += "/" + part
        if path_built in created_cache:
            continue
        try:
            ftp.mkd(path_built)
        except error_perm:
            # existiert schon (550) -> ok; alles andere ist ein echter Fehler
            pass
        created_cache.add(path_built)
    created_cache.add(remote_dir)

def upload_folder_to_ftp_with_progress(local_dir, ftp_dir, allowed_names=None):
    # Dateien sammeln. Optional nur die ausdrücklich freigegebenen Dateien
    # hochladen, damit lokale Hilfs- und ZIP-Dateien nicht auf dem Server landen.
    all_files = []
    for root, _, files in os.walk(local_dir):
        for file in files:
            if allowed_names is not None and file not in allowed_names:
                continue
            local_path = os.path.join(root, file)
            rel_path = os.path.relpath(local_path, local_dir).replace("\\", "/")
            remote_path = os.path.join(ftp_dir, rel_path).replace("\\", "/")
            all_files.append((local_path, remote_path))

    total = len(all_files)
    if total == 0:
        st.warning("Keine Dateien zum Hochladen gefunden.")
        return

    progress_bar = st.progress(0)
    status_text = st.empty()
    t0 = time.perf_counter()

    # 1) Alle Zielverzeichnisse EINMAL vorab anlegen (eine Verbindung).
    #    Danach müssen die Worker nur noch STOR machen, kein mkd.
    status_text.info("Lege Verzeichnisse an ...")
    setup = _new_ftp()
    created = set()
    try:
        for _, remote_path in all_files:
            ensure_ftp_dirs(setup, os.path.dirname(remote_path), created)
    finally:
        try:
            setup.quit()
        except (*ftplib.all_errors, OSError):
            setup.close()

    # 2) Dateien parallel hochladen. Jeder Worker-Thread hält seine eigene
    #    FTP-Verbindung (thread-local) und nutzt sie für alle seine Dateien wieder.
    tl = threading.local()
    conns = []
    conns_lock = threading.Lock()

    def get_conn():
        ftp = getattr(tl, "ftp", None)
        if ftp is None:
            ftp = _new_ftp()
            tl.ftp = ftp
            tl.cwd = None
            with conns_lock:
                conns.append(ftp)
        return ftp

    def upload_one(item):
        local_path, remote_path = item
        ftp = get_conn()
        remote_dir = os.path.dirname(remote_path).replace("\\", "/")
        # cwd pro Verbindung nur wechseln, wenn nötig
        if tl.cwd != remote_dir:
            ftp.cwd(remote_dir)
            tl.cwd = remote_dir
        with open(local_path, "rb") as f:
            ftp.storbinary(f"STOR {os.path.basename(remote_path)}", f)

    workers = max(1, min(FTP_PARALLEL, total))
    uploaded = 0
    errors = []

    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {ex.submit(upload_one, it): it for it in all_files}
        for fut in as_completed(futures):
            local_path, remote_path = futures[fut]
            try:
                fut.result()
            except Exception as e:
                errors.append((os.path.basename(remote_path), str(e)))
            uploaded += 1
            progress_bar.progress(uploaded / total)
            status_text.info(f"Hochgeladen: {uploaded}/{total}")

    # 3) Alle Worker-Verbindungen schließen
    for ftp in conns:
        try:
            ftp.quit()
        except (*ftplib.all_errors, OSError):
            try:
                ftp.close()
            except Exception:
                pass

    dauer = time.perf_counter() - t0
    ok = total - len(errors)
    if errors:
        status_text.warning(
            f"{ok}/{total} hochgeladen, {len(errors)} Fehler "
            f"in {dauer:.1f}s ({workers} parallele Verbindungen)."
        )
        st.code("\n".join(f"{n}: {e}" for n, e in errors[:10]))
    else:
        status_text.success(
            f"Alle {total} Dateien in {dauer:.1f}s hochgeladen "
            f"({workers} parallele Verbindungen, Ø {dauer / total:.2f}s/Datei)."
        )


CSV_COLUMNS = [
    "kw", "fahrer_key", "fahrer_name", "datum", "wochentag",
    "uhrzeit", "tour", "reihenfolge", "url",
]



def _url_mit_cachebuster(url):
    """Ergänzt einen Zeitstempel, damit kein alter Web-Cache ausgeliefert wird."""
    parts = urlsplit(url)
    query = dict(parse_qsl(parts.query, keep_blank_values=True))
    query["_dienstplan_ts"] = str(int(time.time()))
    return urlunsplit((
        parts.scheme,
        parts.netloc,
        parts.path,
        urlencode(query),
        parts.fragment,
    ))


def _csv_datei_sieht_gueltig_aus(path):
    """Verhindert, dass versehentlich eine HTML-Fehler- oder Loginseite übernommen wird."""
    try:
        with open(path, "rb") as handle:
            head = handle.read(8192)
    except OSError:
        return False, "Die heruntergeladene Datei konnte nicht gelesen werden."

    if not head:
        return False, "Die heruntergeladene Datei ist leer."

    text_head = head.decode("utf-8-sig", errors="replace").lstrip()

    if text_head.lower().startswith(("<!doctype html", "<html", "<?xml")):
        return False, "Die Webadresse lieferte HTML statt der CSV-Datei."

    first_line = text_head.splitlines()[0] if text_head.splitlines() else ""
    required = ("kw", "fahrer_key", "fahrer_name", "datum")

    if ";" not in first_line or not all(name in first_line for name in required):
        return (
            False,
            "Die heruntergeladene Datei besitzt nicht den erwarteten "
            "Dienstplan-CSV-Kopf.",
        )

    return True, ""


def download_existing_csv_from_http(local_path):
    """Lädt die bestehende Gesamt-CSV über HTTPS statt über den FTP-Datenkanal."""
    if not DIENSTPLAN_CSV_URL:
        return "not_configured", "Keine DIENSTPLAN_CSV_URL konfiguriert."

    part_path = str(local_path) + ".http.part"
    url = _url_mit_cachebuster(DIENSTPLAN_CSV_URL)
    start_time = time.monotonic()
    downloaded = 0

    try:
        for path in (local_path, part_path):
            try:
                if os.path.exists(path):
                    os.remove(path)
            except OSError:
                pass

        request = Request(
            url,
            headers={
                "User-Agent": "Dienstplan-Upload/1.0",
                "Accept": "text/csv,text/plain;q=0.9,*/*;q=0.1",
                "Cache-Control": "no-cache",
                "Pragma": "no-cache",
            },
            method="GET",
        )

        with urlopen(request, timeout=HTTP_DOWNLOAD_TIMEOUT) as response:
            status_code = getattr(response, "status", 200)
            content_type = (response.headers.get("Content-Type") or "").lower()
            content_length = response.headers.get("Content-Length")

            if status_code != 200:
                return "error", f"HTTPS-Abruf antwortete mit Status {status_code}."

            with open(part_path, "wb") as handle:
                while True:
                    block = response.read(128 * 1024)
                    if not block:
                        break

                    handle.write(block)
                    downloaded += len(block)

                    if downloaded > 100 * 1024 * 1024:
                        raise RuntimeError(
                            "Der HTTPS-Download ist unerwartet größer als 100 MB."
                        )

        if not os.path.isfile(part_path) or os.path.getsize(part_path) <= 0:
            return "error", "Die über HTTPS geladene Datei ist leer."

        valid, validation_message = _csv_datei_sieht_gueltig_aus(part_path)
        if not valid:
            return "error", validation_message

        actual_size = os.path.getsize(part_path)

        if content_length:
            try:
                expected_size = int(content_length)
            except (TypeError, ValueError):
                expected_size = None

            if expected_size and actual_size != expected_size:
                return (
                    "error",
                    "Der HTTPS-Download war unvollständig "
                    f"({actual_size:,} von {expected_size:,} Bytes)."
                    .replace(",", "."),
                )

        os.replace(part_path, local_path)
        elapsed = time.monotonic() - start_time

        return (
            "found",
            "Vorhandene CSV über HTTPS geladen: "
            f"{actual_size / 1024 / 1024:.2f} MB in {elapsed:.1f} Sekunden.",
        )

    except HTTPError as exc:
        if exc.code == 404:
            return (
                "missing",
                f"Unter der Webadresse wurde keine dienstplaene.csv gefunden "
                f"(HTTP 404): {DIENSTPLAN_CSV_URL}",
            )
        if exc.code in (401, 403):
            return (
                "error",
                f"Der Webserver verweigert den Zugriff auf die CSV "
                f"(HTTP {exc.code}).",
            )
        return "error", f"HTTPS-Fehler {exc.code}: {exc.reason}"

    except (URLError, socket.timeout, TimeoutError) as exc:
        return (
            "error",
            f"Zeitüberschreitung oder Verbindungsfehler beim HTTPS-Abruf: {exc}",
        )

    except Exception as exc:
        return "error", f"Vorhandene CSV konnte nicht über HTTPS geladen werden: {exc}"

    finally:
        try:
            if os.path.exists(part_path):
                os.remove(part_path)
        except OSError:
            pass


def download_existing_csv(local_path):
    """Lädt die vorhandene Gesamt-CSV ausschließlich über HTTPS.

    FTP bleibt nur für den späteren Upload zuständig. Dadurch wird verhindert,
    dass blockierte passive oder aktive FTP-Datenkanäle die Verarbeitung
    minutenlang aufhalten.
    """
    if not DIENSTPLAN_CSV_URL:
        return (
            "error",
            "DIENSTPLAN_CSV_URL fehlt in der .env. "
            "Beispiel: https://DEINE-DOMAIN/test/uploads/2026/dienstplaene.csv",
        )

    st.info(f"Lade die vorhandene CSV über HTTPS: {DIENSTPLAN_CSV_URL}")
    return download_existing_csv_from_http(local_path)


def download_existing_csv_from_ftp(local_path):
    """Vorhandene dienstplaene.csv eindeutig prüfen und robust laden.

    Zuerst wird die Datei über den FTP-Steuerkanal mit SIZE geprüft. Fehlt sie,
    kommt sofort der Status "missing". Ist sie vorhanden, wird zunächst im
    passiven Modus geladen. Blockiert der Datenkanal, erfolgt automatisch ein
    zweiter Versuch im aktiven Modus.
    """
    part_path = str(local_path) + ".part"
    remote_dir = (FTP_BASE_DIR or "/").replace("\\", "/").rstrip("/") or "/"
    remote_display = (
        f"{remote_dir}/dienstplaene.csv"
        if remote_dir != "/"
        else "/dienstplaene.csv"
    )

    for path in (local_path, part_path):
        try:
            if os.path.exists(path):
                os.remove(path)
        except OSError:
            pass

    def connect_ftp(passive_mode):
        ftp = _new_ftp()
        ftp.set_pasv(passive_mode)

        if getattr(ftp, "sock", None) is not None:
            ftp.sock.settimeout(FTP_TIMEOUT)

        if remote_dir != "/":
            ftp.cwd(remote_dir)

        # SIZE funktioniert bei vielen Servern nur im Binärmodus.
        try:
            ftp.voidcmd("TYPE I")
        except ftplib.all_errors:
            pass

        return ftp

    def close_ftp(ftp):
        if ftp is None:
            return
        try:
            ftp.close()
        except Exception:
            pass

    # 1) Existenz und Größe ausschließlich über den Steuerkanal prüfen.
    probe = None
    remote_size = None
    try:
        probe = connect_ftp(True)
        try:
            remote_size = probe.size("dienstplaene.csv")
        except error_perm as exc:
            if str(exc).lstrip().startswith("550"):
                return (
                    "missing",
                    f"Nicht gefunden: {remote_display}. "
                    "Bitte FTP_BASE_DIR und den exakten Dateinamen prüfen.",
                )
            raise

        if remote_size is None:
            # Manche Server unterstützen SIZE nicht. Mit MDTM wird wenigstens
            # geprüft, ob die Datei angesprochen werden kann.
            try:
                probe.sendcmd("MDTM dienstplaene.csv")
            except error_perm as exc:
                if str(exc).lstrip().startswith("550"):
                    return (
                        "missing",
                        f"Nicht gefunden: {remote_display}. "
                        "Bitte FTP_BASE_DIR und den exakten Dateinamen prüfen.",
                    )
                raise

        size_text = (
            f"{int(remote_size) / 1024 / 1024:.2f} MB"
            if remote_size is not None
            else "Größe unbekannt"
        )
        st.info(f"Datei gefunden: {remote_display} ({size_text}). Starte Download …")

    except error_perm as exc:
        return "error", f"FTP-Zugriffsfehler bei {remote_display}: {exc}"
    except (socket.timeout, TimeoutError) as exc:
        return (
            "error",
            f"Schon die Prüfung von {remote_display} ist abgelaufen: {exc}",
        )
    except Exception as exc:
        return (
            "error",
            f"Die Datei konnte unter {remote_display} nicht geprüft werden: {exc}",
        )
    finally:
        close_ftp(probe)

    # 2) Download zuerst passiv, danach als automatische Alternative aktiv.
    errors = []

    for passive_mode, mode_name in ((True, "passiv"), (False, "aktiv")):
        ftp = None
        downloaded = 0
        start_time = time.monotonic()

        try:
            try:
                if os.path.exists(part_path):
                    os.remove(part_path)
            except OSError:
                pass

            st.info(f"FTP-Download im {mode_name}en Modus …")
            ftp = connect_ftp(passive_mode)

            with open(part_path, "wb") as handle:
                def write_block(block):
                    nonlocal downloaded

                    elapsed = time.monotonic() - start_time
                    if elapsed > FTP_DOWNLOAD_MAX_SECONDS:
                        raise TimeoutError(
                            f"{mode_name.capitalize()}er Download nach "
                            f"{FTP_DOWNLOAD_MAX_SECONDS} Sekunden abgebrochen."
                        )

                    handle.write(block)
                    downloaded += len(block)

                ftp.retrbinary(
                    "RETR dienstplaene.csv",
                    write_block,
                    blocksize=128 * 1024,
                )

            elapsed = time.monotonic() - start_time
            actual_size = os.path.getsize(part_path) if os.path.isfile(part_path) else 0

            if actual_size <= 0:
                raise RuntimeError("Die heruntergeladene Datei ist leer.")

            if (
                remote_size is not None
                and int(remote_size) > 0
                and actual_size != int(remote_size)
            ):
                raise RuntimeError(
                    "Der Download war unvollständig "
                    f"({actual_size:,} von {int(remote_size):,} Bytes)."
                    .replace(",", ".")
                )

            os.replace(part_path, local_path)

            return (
                "found",
                f"Vorhandene CSV im {mode_name}en FTP-Modus geladen: "
                f"{actual_size / 1024 / 1024:.2f} MB in {elapsed:.1f} Sekunden.",
            )

        except (socket.timeout, TimeoutError) as exc:
            errors.append(f"{mode_name}: Zeitüberschreitung ({exc})")
        except error_perm as exc:
            if str(exc).lstrip().startswith("550"):
                return (
                    "missing",
                    f"Die zuvor gefundene Datei ist nicht mehr vorhanden: {remote_display}.",
                )
            errors.append(f"{mode_name}: FTP-Fehler ({exc})")
        except Exception as exc:
            errors.append(f"{mode_name}: {exc}")
        finally:
            close_ftp(ftp)
            try:
                if os.path.exists(part_path):
                    os.remove(part_path)
            except OSError:
                pass

    return (
        "error",
        "Die Datei wurde gefunden, aber beide FTP-Datenmodi sind gescheitert. "
        + " | ".join(errors),
    )

def read_dienstplan_csv(path):
    """Dienstplan-CSV robust lesen und auf das erwartete Schema bringen."""
    last_error = None
    dataframe = None

    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            dataframe = pd.read_csv(
                path,
                sep=";",
                dtype=str,
                encoding=encoding,
                keep_default_na=False,
            )
            break
        except Exception as exc:
            last_error = exc

    if dataframe is None:
        raise ValueError(f"Vorhandene dienstplaene.csv ist nicht lesbar: {last_error}")

    dataframe.columns = [str(column).lstrip("\ufeff").strip() for column in dataframe.columns]

    required = {"kw", "fahrer_key", "fahrer_name", "datum", "uhrzeit", "tour"}
    missing = sorted(required.difference(dataframe.columns))
    if missing:
        raise ValueError(
            "Vorhandene dienstplaene.csv hat nicht das erwartete Format. "
            f"Fehlende Spalten: {', '.join(missing)}"
        )

    for column in CSV_COLUMNS:
        if column not in dataframe.columns:
            dataframe[column] = ""

    return dataframe[CSV_COLUMNS].copy()


def week_start_sunday_from_dates(values):
    """Sonntag der jeweiligen Planwoche für Datumswerte ermitteln."""
    dates = pd.to_datetime(values, errors="coerce")
    weekday = dates.dt.weekday
    return dates - pd.to_timedelta((weekday + 1) % 7, unit="D")


def clean_and_sort_csv(dataframe):
    """Doppelte/überflüssige Zeilen entfernen und Reihenfolgen neu bilden."""
    if dataframe.empty:
        return pd.DataFrame(columns=CSV_COLUMNS)

    dataframe = dataframe.copy()
    for column in CSV_COLUMNS:
        if column not in dataframe.columns:
            dataframe[column] = ""
    dataframe = dataframe[CSV_COLUMNS]

    dataframe["kw"] = pd.to_numeric(dataframe["kw"], errors="coerce")
    dataframe = dataframe.loc[dataframe["kw"].notna()].copy()
    dataframe["kw"] = dataframe["kw"].astype(int)

    dataframe = dataframe.drop_duplicates(
        subset=[
            "kw", "fahrer_key", "fahrer_name", "datum",
            "wochentag", "uhrzeit", "tour"
        ],
        keep="first",
    ).copy()

    ist_leer = (
        dataframe["uhrzeit"].astype(str).str.strip().isin(["", "–"])
        & dataframe["tour"].astype(str).str.strip().isin(["", "–"])
    )
    hat_echten_eintrag = (~ist_leer).groupby([
        dataframe["kw"],
        dataframe["fahrer_key"],
        dataframe["datum"],
    ]).transform("any")
    dataframe = dataframe.loc[~(ist_leer & hat_echten_eintrag)].copy()

    dataframe["_datum_sort"] = pd.to_datetime(dataframe["datum"], errors="coerce")
    dataframe["_wochenstart"] = week_start_sunday_from_dates(dataframe["datum"])
    dataframe["_reihenfolge_alt"] = pd.to_numeric(
        dataframe["reihenfolge"], errors="coerce"
    ).fillna(9999)

    dataframe = dataframe.sort_values(
        by=[
            "_wochenstart", "fahrer_name", "_datum_sort",
            "_reihenfolge_alt"
        ],
        kind="stable",
    )
    dataframe["reihenfolge"] = (
        dataframe.groupby(["_wochenstart", "fahrer_key"], dropna=False).cumcount() + 1
    )
    dataframe = dataframe.drop(
        columns=["_datum_sort", "_wochenstart", "_reihenfolge_alt"]
    )

    return dataframe[CSV_COLUMNS]


def merge_uploaded_weeks(existing_df, new_df, uploaded_week_starts):
    """Nur die neu hochgeladenen Sonntag-bis-Samstag-Wochen ersetzen."""
    if existing_df is None or existing_df.empty:
        return clean_and_sort_csv(new_df), 0, 0

    existing_df = existing_df.copy()
    existing_dates = pd.to_datetime(existing_df["datum"], errors="coerce").dt.normalize()
    replace_mask = pd.Series(False, index=existing_df.index)

    for week_start in sorted(uploaded_week_starts):
        start = pd.Timestamp(week_start).normalize()
        end = start + pd.Timedelta(days=6)
        replace_mask |= existing_dates.between(start, end, inclusive="both")

    removed = int(replace_mask.sum())
    retained_df = existing_df.loc[~replace_mask].copy()
    retained = len(retained_df)

    merged = pd.concat([retained_df, new_df], ignore_index=True)
    return clean_and_sort_csv(merged), retained, removed


def normalize_driver_name(nachname, vorname):
    """Fahrernamen bereinigen; leere Excel-Zellen können als 0/0.0 erscheinen."""
    def clean_part(value):
        if pd.isna(value):
            return ""

        value_str = str(value).strip()
        value_lower = value_str.lower()

        if value_lower in {"", "nan", "none", "null", "nat", "-", "–"}:
            return ""
        if re.fullmatch(r"0+(?:[.,]0+)?", value_str):
            return ""

        return value_str.title()

    nachname = clean_part(nachname)
    vorname = clean_part(vorname)

    # Bekannte Schreibvarianten aus unterschiedlichen Excel-Dateien
    # auf eine einheitliche Fahreridentität zusammenführen.
    alias = {
        ("khalleefah", ""): ("Khalleefah", "Saed Awami Sayid"),
        ("alem", "mohamed"): ("Alem", "Mohammed"),
        ("maghraoui", "zakaria"): ("Maghraoui", "Zakariae"),
    }.get((nachname.casefold(), vorname.casefold()))
    if alias:
        nachname, vorname = alias

    if nachname and vorname:
        return f"{nachname}, {vorname}"
    if nachname:
        return nachname
    if vorname:
        return vorname
    return ""

def parse_uhrzeit(uhrzeit):
    """Excel-Uhrzeiten schnell und ohne teure Einzel-Konvertierungen lesen."""
    if uhrzeit is None or pd.isna(uhrzeit):
        return "–"

    if isinstance(uhrzeit, datetime):
        return uhrzeit.strftime("%H:%M")

    if isinstance(uhrzeit, datetime_time):
        return uhrzeit.strftime("%H:%M")

    if isinstance(uhrzeit, (int, float)):
        wert = float(uhrzeit)
        if wert == 0:
            return "00:00"
        # Excel speichert Uhrzeiten normalerweise als Tagesbruchteil.
        if 0 <= wert < 1:
            minuten = int(round(wert * 24 * 60)) % (24 * 60)
            return f"{minuten // 60:02d}:{minuten % 60:02d}"
        return str(uhrzeit).strip()

    uhrzeit_str = str(uhrzeit).strip()
    if not uhrzeit_str or uhrzeit_str.lower() in {"nan", "none", "nat"}:
        return "–"

    match = re.match(r"^(\d{1,2}):(\d{2})", uhrzeit_str)
    if match:
        return f"{int(match.group(1)):02d}:{match.group(2)}"

    return uhrzeit_str

def parse_tour(tour):
    if pd.isna(tour):
        return "–"

    tour_str = str(tour).strip()
    if not tour_str or tour_str.lower() == "nan":
        return "–"

    return tour_str

def generate_shared_html(css_styles):
    """Erzeugt eine CSP-konforme HTML-Datei ohne eingebettetes JavaScript."""
    template = r'''<!DOCTYPE html>
<html lang="de">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
  <meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
  <meta http-equiv="Pragma" content="no-cache">
  <meta http-equiv="Expires" content="0">
  <title>Dienstplan</title>
  <style>__CSS_STYLES__</style>
</head>
<body>
<div class="container-outer">
  <div class="back-bar">
    <a href="../../../plane.php" class="btn-back" id="btnBack">
      <span class="btn-back-arrow" aria-hidden="true">‹</span>
      <span>Zurück</span>
    </a>
  </div>

  <header class="plan-hero">
    <div class="hero-topline">
      <span class="hero-dot"></span>
      <span>Dienstplan</span>
    </div>

    <div class="hero-main">
      <div class="hero-kw">
        <span class="hero-kw-label">Woche</span>
        <span class="hero-kw-number" id="kwNumber">--</span>
      </div>

      <div class="hero-meta">
        <div class="headline-name" id="driverName">Dienstplan wird geladen</div>
        <div class="headline-period" id="periodText">Bitte einen Moment</div>
      </div>
    </div>
  </header>

  <main class="week-list" id="weekList"></main>
</div>
<div class="browser-safe-spacer" aria-hidden="true"></div>

<script src="dienstplan_app_v4.js?v=1" defer></script>

</body>
</html>'''
    return template.replace("__CSS_STYLES__", css_styles)


def generate_shared_js():
    """JavaScript der gemeinsamen Dienstplanseite als externe Datei."""
    return r'''(function () {
  "use strict";

  const params = new URLSearchParams(window.location.search);
  const requestedKw = params.get("kw");
  const requestedDriver = params.get("fahrer") || params.get("driver");

  const weekList = document.getElementById("weekList");
  const kwNumber = document.getElementById("kwNumber");
  const driverName = document.getElementById("driverName");
  const periodText = document.getElementById("periodText");

  function escapeHtml(value) {
    return String(value ?? "")
      .replaceAll("&", "&amp;")
      .replaceAll("<", "&lt;")
      .replaceAll(">", "&gt;")
      .replaceAll('"', "&quot;")
      .replaceAll("'", "&#039;");
  }

  function parseCsv(text, delimiter = ";") {
    const rows = [];
    let row = [];
    let field = "";
    let quoted = false;

    for (let i = 0; i < text.length; i += 1) {
      const char = text[i];

      if (quoted) {
        if (char === '"') {
          if (text[i + 1] === '"') {
            field += '"';
            i += 1;
          } else {
            quoted = false;
          }
        } else {
          field += char;
        }
      } else if (char === '"') {
        quoted = true;
      } else if (char === delimiter) {
        row.push(field);
        field = "";
      } else if (char === "\n") {
        row.push(field.replace(/\r$/, ""));
        rows.push(row);
        row = [];
        field = "";
      } else {
        field += char;
      }
    }

    if (field.length > 0 || row.length > 0) {
      row.push(field.replace(/\r$/, ""));
      rows.push(row);
    }

    if (!rows.length) return [];

    const headers = rows.shift().map((header, index) =>
      index === 0 ? header.replace(/^\uFEFF/, "") : header
    );

    return rows
      .filter(values => values.some(value => value !== ""))
      .map(values => {
        const item = {};
        headers.forEach((header, index) => {
          item[header] = values[index] ?? "";
        });
        return item;
      });
  }

  function normalizeKw(value) {
    const number = Number.parseInt(value, 10);
    return Number.isFinite(number) ? String(number) : "";
  }

  function formatDate(isoDate) {
    const parts = String(isoDate).split("-");
    if (parts.length !== 3) return isoDate;
    return `${parts[2]}.${parts[1]}.${parts[0]}`;
  }

  function createDayCard(entry) {
    const weekday = entry.wochentag || "";
    const time = entry.uhrzeit || "–";
    const tour = entry.tour || "–";
    const contentCheck = `${time} ${tour}`.toLowerCase();

    let cardClass = "daycard";
    let badgeText = "Dienst";
    let icon = "🚚";

    if (weekday === "Samstag") {
      cardClass += " samstag";
      badgeText = "Samstag";
    } else if (weekday === "Sonntag") {
      cardClass += " sonntag";
      badgeText = "Sonntag";
    }

    if (tour === "–" && time === "–") {
      cardClass += " leer";
      badgeText = "Kein Eintrag";
      icon = "—";
    } else if (contentCheck.includes("urlaub")) {
      cardClass += " frei";
      badgeText = "Urlaub";
      icon = "☀️";
    } else if (contentCheck.includes("frei")) {
      cardClass += " frei";
      badgeText = "Frei";
      icon = "🕊️";
    } else if (contentCheck.includes("ausgleich")) {
      cardClass += " frei";
      badgeText = "Ausgleich";
      icon = "🛌";
    } else if (contentCheck.includes("krank")) {
      cardClass += " krank";
      badgeText = "Krank";
      icon = "💊";
    }

    return `
      <section class="${cardClass}">
        <div class="day-top">
          <div class="day-date">
            <div class="weekday">${escapeHtml(weekday)}</div>
            <div class="prominent-date">${escapeHtml(formatDate(entry.datum))}</div>
          </div>
          <div class="day-badge"><span>${icon}</span>${escapeHtml(badgeText)}</div>
        </div>

        <div class="info">
          <div class="info-block tour-block">
            <span class="label">Tour / Aufgabe:</span>
            <span class="value">${escapeHtml(tour)}</span>
          </div>
          <div class="info-block time-block">
            <span class="label">Uhrzeit:</span>
            <span class="value">${escapeHtml(time)}</span>
          </div>
        </div>
      </section>`;
  }

  function showMessage(title, detail) {
    driverName.textContent = title;
    periodText.textContent = detail;
    weekList.innerHTML = `
      <section class="daycard leer">
        <div class="info-block">
          <span class="value">${escapeHtml(detail)}</span>
        </div>
      </section>`;
  }

  if (!requestedKw || !requestedDriver) {
    showMessage(
      "Dienstplan nicht ausgewählt",
      "Der Aufruf benötigt die Angaben kw und fahrer, zum Beispiel: dienstplan_v4.html?kw=26&fahrer=Mueller"
    );
    return;
  }

  fetch(`dienstplaene.csv?v=${Date.now()}`, { cache: "no-store" })
    .then(response => {
      if (!response.ok) {
        throw new Error(`CSV konnte nicht geladen werden (${response.status}).`);
      }
      return response.text();
    })
    .then(text => {
      const data = parseCsv(text);
      const selected = data
        .filter(entry =>
          normalizeKw(entry.kw) === normalizeKw(requestedKw) &&
          entry.fahrer_key === requestedDriver
        )
        .sort((a, b) => Number(a.reihenfolge) - Number(b.reihenfolge));

      if (!selected.length) {
        kwNumber.textContent = String(requestedKw).padStart(2, "0");
        showMessage(
          "Kein Dienstplan gefunden",
          `Für Kalenderwoche ${requestedKw} und Fahrer ${requestedDriver} sind keine Daten vorhanden.`
        );
        return;
      }

      const first = selected[0];
      const dates = selected.map(entry => entry.datum).filter(Boolean).sort();
      const startDate = dates[0];
      const endDate = dates[dates.length - 1];

      kwNumber.textContent = String(first.kw).padStart(2, "0");
      driverName.textContent = first.fahrer_name;
      periodText.textContent = `${formatDate(startDate)} – ${formatDate(endDate)}`;
      document.title = `KW${String(first.kw).padStart(2, "0")} – ${first.fahrer_name}`;
      weekList.innerHTML = selected.map(createDayCard).join("");
    })
    .catch(error => {
      showMessage("Fehler beim Laden", error.message);
    });
})();
'''


css_styles = """
:root {
  --bg-main: #e6eef7;
  --card: #ffffff;
  --border: #d7e0ec;
  --border-strong: #b8c6d8;
  --text: #0f172a;
  --muted: #64748b;
  --soft: #f8fafc;
  --soft-blue: #eef6ff;
  --blue: #1b66b3;
  --blue-dark: #1b3a7a;
  --green: #15803d;
  --green-soft: #ecfdf3;
  --yellow: #b45309;
  --yellow-soft: #fffbeb;
  --red: #b91c1c;
  --red-soft: #fff1f2;
  --shadow: 0 10px 24px rgba(15, 23, 42, .08);
  --shadow-soft: 0 2px 8px rgba(15, 23, 42, .06);
}

* {
  box-sizing: border-box;
}

html {
  -webkit-text-size-adjust: 100%;
  min-height: 100%;
  min-height: -webkit-fill-available;
  min-height: 100svh;
  min-height: 100dvh;
}

body {
  margin: 0;
  padding: 0;
  min-height: 100%;
  min-height: -webkit-fill-available;
  min-height: calc(100svh + 2px);
  min-height: calc(100dvh + 2px);
  padding-bottom: calc(150px + env(safe-area-inset-bottom) + var(--browser-bottom-gap, 0px));
  background:
    radial-gradient(circle at top left, rgba(27, 102, 179, .10), transparent 32rem),
    var(--bg-main);
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif;
  color: var(--text);
  font-size: 15px;
  line-height: 1.45;
  overflow-x: hidden;
}

.container-outer {
  width: min(560px, calc(100vw - 24px));
  margin: 16px auto 0;
  padding-bottom: calc(80px + env(safe-area-inset-bottom) + var(--browser-bottom-gap, 0px));
}

.browser-safe-spacer {
  height: calc(140px + env(safe-area-inset-bottom) + var(--browser-bottom-gap, 0px));
  pointer-events: none;
}

.back-bar {
  margin-bottom: 10px;
}

.btn-back {
  display: inline-flex;
  align-items: center;
  gap: 4px;
  font-size: .82rem;
  font-weight: 700;
  text-decoration: none;
  color: var(--blue-dark);
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: 999px;
  padding: 7px 14px 7px 11px;
  box-shadow: var(--shadow-soft);
  -webkit-tap-highlight-color: transparent;
  transition: background .15s ease, border-color .15s ease, transform .05s ease;
}

.btn-back:hover {
  background: var(--soft-blue);
  border-color: var(--border-strong);
}

.btn-back:active {
  transform: scale(.97);
}

.btn-back-arrow {
  font-size: 1.15rem;
  line-height: 1;
  margin-top: -1px;
}

@media print {
  .back-bar {
    display: none;
  }
}

.plan-hero {
  background: linear-gradient(135deg, #ffffff 0%, #eef6ff 100%);
  border: 1px solid var(--border);
  border-radius: 18px;
  padding: 14px;
  box-shadow: var(--shadow);
  margin-bottom: 10px;
  overflow: hidden;
  position: relative;
}

.plan-hero::after {
  content: "";
  position: absolute;
  width: 90px;
  height: 90px;
  right: -34px;
  top: -36px;
  border-radius: 999px;
  background: rgba(27, 102, 179, .12);
}

.hero-topline {
  display: inline-flex;
  align-items: center;
  gap: 6px;
  font-size: .70rem;
  font-weight: 800;
  letter-spacing: .05em;
  text-transform: uppercase;
  color: var(--blue-dark);
  background: rgba(27, 102, 179, .08);
  border: 1px solid rgba(27, 102, 179, .12);
  border-radius: 999px;
  padding: 4px 9px;
  margin-bottom: 10px;
  position: relative;
  z-index: 1;
}

.hero-dot {
  width: 7px;
  height: 7px;
  border-radius: 50%;
  background: var(--blue);
  display: inline-block;
}

.hero-main {
  display: grid;
  grid-template-columns: auto 1fr;
  gap: 12px;
  align-items: center;
  position: relative;
  z-index: 1;
}

.hero-kw {
  width: 76px;
  min-height: 70px;
  border-radius: 16px;
  background: linear-gradient(135deg, var(--blue), var(--blue-dark));
  color: #ffffff;
  display: flex;
  flex-direction: column;
  justify-content: center;
  align-items: center;
  box-shadow: 0 8px 18px rgba(27, 102, 179, .25);
}

.hero-kw-label {
  font-size: .52rem;
  font-weight: 800;
  text-transform: uppercase;
  letter-spacing: .05em;
  opacity: .78;
}

.hero-kw-number {
  font-size: 1.85rem;
  line-height: 1;
  font-weight: 900;
  letter-spacing: -.03em;
}

.hero-meta {
  min-width: 0;
}

.headline-name {
  font-size: 1.08rem;
  line-height: 1.15;
  font-weight: 900;
  color: var(--text);
  overflow-wrap: anywhere;
}

.headline-period {
  margin-top: 4px;
  font-size: .82rem;
  font-weight: 700;
  color: var(--muted);
}

.week-list {
  display: flex;
  flex-direction: column;
  gap: 8px;
}

.daycard {
  background: rgba(255,255,255,.92);
  border: 1px solid var(--border);
  border-radius: 16px;
  padding: 10px;
  box-shadow: var(--shadow-soft);
  transition: transform .12s ease, box-shadow .12s ease, border-color .12s ease;
}

.daycard:hover {
  transform: translateY(-1px);
  box-shadow: 0 8px 18px rgba(15, 23, 42, .10);
  border-color: var(--border-strong);
}

.day-top {
  display: flex;
  justify-content: space-between;
  align-items: center;
  gap: 10px;
  margin-bottom: 8px;
}

.day-date {
  min-width: 0;
}

.weekday {
  font-size: .78rem;
  font-weight: 900;
  color: var(--blue-dark);
  line-height: 1.1;
}

.prominent-date {
  font-size: .72rem;
  color: var(--muted);
  font-weight: 700;
  margin-top: 2px;
}

.day-badge {
  display: inline-flex;
  align-items: center;
  gap: 5px;
  flex: 0 0 auto;
  font-size: .66rem;
  font-weight: 900;
  color: var(--blue-dark);
  background: var(--soft-blue);
  border: 1px solid #dbeafe;
  border-radius: 999px;
  padding: 4px 8px;
}

.info {
  display: grid;
  grid-template-columns: minmax(0, 1fr) 118px;
  gap: 7px;
}

.info-block {
  min-width: 0;
  display: flex;
  flex-direction: column;
  gap: 2px;
  background: var(--soft);
  border: 1px solid #e2e8f0;
  border-radius: 12px;
  padding: 8px 9px;
}

.label {
  color: #94a3b8;
  font-size: .58rem;
  font-weight: 900;
  text-transform: uppercase;
  letter-spacing: .045em;
  white-space: nowrap;
}

.value {
  color: var(--text);
  font-size: .88rem;
  font-weight: 900;
  line-height: 1.2;
  overflow-wrap: anywhere;
}

.time-block .value {
  white-space: nowrap;
}

.daycard.samstag,
.daycard.sonntag {
  background: linear-gradient(180deg, #ffffff, var(--yellow-soft));
  border-color: #fde68a;
}

.daycard.samstag .day-badge,
.daycard.sonntag .day-badge {
  color: var(--yellow);
  background: #fef3c7;
  border-color: #fde68a;
}

.daycard.samstag .weekday,
.daycard.sonntag .weekday {
  color: var(--yellow);
}

.daycard.frei {
  background: linear-gradient(180deg, #ffffff, var(--green-soft));
  border-color: #bbf7d0;
}

.daycard.frei .day-badge {
  color: var(--green);
  background: #dcfce7;
  border-color: #bbf7d0;
}

.daycard.frei .weekday {
  color: var(--green);
}

.daycard.krank {
  background: linear-gradient(180deg, #ffffff, var(--red-soft));
  border-color: #fecdd3;
}

.daycard.krank .day-badge {
  color: var(--red);
  background: #ffe4e6;
  border-color: #fecdd3;
}

.daycard.krank .weekday {
  color: var(--red);
}

.daycard.leer {
  opacity: .78;
}

.daycard.leer .value {
  color: #94a3b8;
}

@media (max-width: 440px) {
  .container-outer {
    width: min(100vw - 16px, 560px);
    margin-top: 8px;
    margin-bottom: 0;
  }

  .plan-hero {
    border-radius: 16px;
    padding: 12px;
  }

  .hero-main {
    grid-template-columns: 74px 1fr;
    gap: 10px;
  }

  .hero-kw {
    width: 74px;
    min-height: 64px;
    border-radius: 14px;
  }

  .hero-kw-label {
    font-size: .54rem;
  }

  .hero-kw-number {
    font-size: 1.65rem;
  }

  .headline-name {
    font-size: 1rem;
  }

  .headline-period {
    font-size: .76rem;
  }

  .daycard {
    border-radius: 14px;
    padding: 9px;
  }

  /* Handy: Tour und Uhrzeit bleiben nebeneinander, solange genug Platz ist. */
  .info {
    grid-template-columns: minmax(0, 1fr) 92px;
    gap: 6px;
  }

  .info-block {
    padding: 7px 7px;
  }

  .label {
    font-size: .50rem;
    letter-spacing: .025em;
  }

  .value {
    font-size: .80rem;
  }

  .time-block .value {
    white-space: nowrap;
  }
}

@media (max-width: 340px) {
  .info {
    grid-template-columns: 1fr;
  }

  .time-block .value {
    white-space: normal;
  }
}

@media print {
  body {
    background: #ffffff;
  }

  .container-outer {
    width: 100%;
    margin: 0;
    padding-bottom: 0;
  }

  .browser-safe-spacer {
    display: none;
  }

  .plan-hero,
  .daycard {
    box-shadow: none;
  }
}
"""


def _ist_leer(value):
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return str(value).strip().lower() in {"", "nan", "none", "null", "nat"}


def _datum_lesen(value):
    """Excel-Datum robust und schnell in ein Python-Datum umwandeln."""
    if _ist_leer(value):
        return None

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            result = from_excel(value)
            if isinstance(result, datetime):
                return result.date()
            if isinstance(result, date):
                return result
        except Exception:
            return None

    try:
        parsed = pd.to_datetime(str(value).strip(), errors="coerce", dayfirst=True)
        if pd.isna(parsed):
            return None
        return parsed.date()
    except Exception:
        return None


def _arbeitsmappe_schnell_lesen(uploaded_file):
    """Nur die tatsächlich benötigten Excel-Spalten einlesen.

    Der alte Pandas-Import konnte bei formatierten Leerzeilen bis ans Ende des
    Excel-Blatts laufen. Hier wird nach längeren Leerbereichen abgebrochen.
    """
    datei_bytes = uploaded_file.getvalue()
    workbook = load_workbook(
        BytesIO(datei_bytes),
        read_only=True,
        data_only=True,
    )

    try:
        benoetigte_blaetter = {"Touren", "a Fahrer"}
        fehlend = sorted(benoetigte_blaetter.difference(workbook.sheetnames))
        if fehlend:
            raise ValueError(
                f"In {uploaded_file.name} fehlen die Blätter: {', '.join(fehlend)}"
            )

        # Fahrerblatt: B = Nachname, C = Vorname, Daten ab Zeile 2.
        fahrer = []
        bekannte_fahrer = set()
        ws_fahrer = workbook["a Fahrer"]
        leere_zeilen = 0

        for nachname, vorname in ws_fahrer.iter_rows(
            min_row=2,
            min_col=2,
            max_col=3,
            values_only=True,
        ):
            fahrer_name = normalize_driver_name(nachname, vorname)
            if not fahrer_name:
                leere_zeilen += 1
                # Ein langer Leerbereich bedeutet in diesen Dateien das Ende.
                if leere_zeilen >= 250:
                    break
                continue

            leere_zeilen = 0
            if fahrer_name not in bekannte_fahrer:
                bekannte_fahrer.add(fahrer_name)
                fahrer.append(fahrer_name)

        # Tourenblatt: Daten ab Zeile 6.
        # D/E = Fahrer 1, G/H = Fahrer 2, I = Uhrzeit, O = Datum, P = Tour.
        ws_touren = workbook["Touren"]
        touren = []
        leere_zeilen = 0

        for values in ws_touren.iter_rows(
            min_row=6,
            min_col=4,
            max_col=16,
            values_only=True,
        ):
            nach1, vor1 = values[0], values[1]
            nach2, vor2 = values[3], values[4]
            uhrzeit = values[5]
            datum_raw = values[11]
            tour = values[12]

            relevante_werte = (nach1, vor1, nach2, vor2, uhrzeit, datum_raw, tour)
            if all(_ist_leer(value) for value in relevante_werte):
                leere_zeilen += 1
                if leere_zeilen >= 500:
                    break
                continue

            leere_zeilen = 0
            datum = _datum_lesen(datum_raw)
            if datum is None:
                continue

            fahrer1 = normalize_driver_name(nach1, vor1)
            fahrer2 = normalize_driver_name(nach2, vor2)
            if not fahrer1 and not fahrer2:
                continue

            touren.append({
                "datum": datum,
                "uhrzeit": parse_uhrzeit(uhrzeit),
                "tour": parse_tour(tour),
                "fahrer": [name for name in (fahrer1, fahrer2) if name],
            })

        return fahrer, touren
    finally:
        workbook.close()


st.set_page_config(page_title="Touren-Export", layout="centered")
st.title("Dienstplan aktualisieren")
st.caption("Die Verarbeitung startet erst nach Klick auf „Dienstpläne verarbeiten“.")

with st.form("dienstplan_upload_form", clear_on_submit=False):
    uploaded_files = st.file_uploader(
        "Excel-Dateien hochladen (Blatt 'Touren' und 'a Fahrer')",
        type=["xlsx"],
        accept_multiple_files=True,
    )
    auto_ftp = st.checkbox("Nach der Verarbeitung automatisch auf FTP hochladen", value=False)
    submitted = st.form_submit_button("Dienstpläne verarbeiten", type="primary")

if submitted:
    if not uploaded_files:
        st.warning("Bitte mindestens eine Excel-Datei auswählen.")
        st.stop()

    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            zip_path = os.path.join(tmpdir, "gesamt_export.zip")
            html_path = os.path.join(tmpdir, "dienstplan_v4.html")
            js_path = os.path.join(tmpdir, "dienstplan_app_v4.js")
            csv_path = os.path.join(tmpdir, "dienstplaene.csv")
            existing_csv_path = os.path.join(tmpdir, "dienstplaene_vorhanden.csv")

            ausschluss_stichwoerter = [
                "zippel", "insel", "paasch", "meyer",
                "ihde", "devies", "insellogistik",
            ]

            sonder_dateien = {
                ("fechner", "klaus"): "KFechner",
                ("fechner", "danny"): "Fechner",
                ("scheil", "rene"): "RScheil",
                ("scheil", "eric"): "Scheil",
                ("schulz", "julian"): "Schulz",
                ("schulz", "stephan"): "STSchulz",
                ("lewandowski", "kamil"): "Lewandowski",
                ("lewandowski", "dominik"): "DLewandowski",
                ("schlutt", "rene"): "Schlutt",
                ("schlutt", "hubert"): "HSchlutt",
            }

            csv_rows = []
            bekannte_zeilen = set()
            fahrer_wochen = set()
            hochgeladene_wochen = set()

            fortschritt = st.progress(0)
            status = st.empty()
            gesamt_start = time.perf_counter()

            for datei_index, uploaded_file in enumerate(uploaded_files, start=1):
                status.info(
                    f"Lese {uploaded_file.name} "
                    f"({datei_index}/{len(uploaded_files)}) …"
                )
                datei_start = time.perf_counter()
                fahrer_liste, touren_liste = _arbeitsmappe_schnell_lesen(uploaded_file)

                if not touren_liste:
                    st.warning(f"In {uploaded_file.name} wurden keine gültigen Tourenzeilen gefunden.")
                    fortschritt.progress(datei_index / len(uploaded_files))
                    continue

                fahrer_dict = {fahrer_name: {} for fahrer_name in fahrer_liste}
                global_start_datum = min(eintrag["datum"] for eintrag in touren_liste)

                for eintrag in touren_liste:
                    tag = eintrag["datum"]
                    eintrag_text = f'{eintrag["uhrzeit"]} – {eintrag["tour"]}'

                    for fahrer_name in eintrag["fahrer"]:
                        fahrer_dict.setdefault(fahrer_name, {})
                        tag_liste = fahrer_dict[fahrer_name].setdefault(tag, [])
                        if eintrag_text not in tag_liste:
                            tag_liste.append(eintrag_text)

                global_start_sonntag = global_start_datum - pd.Timedelta(
                    days=(global_start_datum.weekday() + 1) % 7
                )
                global_kw = get_plan_kw(global_start_sonntag)
                hochgeladene_wochen.add(pd.Timestamp(global_start_sonntag).date())

                for fahrer_name in sorted(fahrer_dict, key=str.casefold):
                    eintraege = fahrer_dict[fahrer_name]
                    if eintraege:
                        start_datum = min(eintraege)
                        start_sonntag = start_datum - pd.Timedelta(
                            days=(start_datum.weekday() + 1) % 7
                        )
                        kw = get_plan_kw(start_sonntag)
                    else:
                        start_sonntag = global_start_sonntag
                        kw = global_kw

                    if "," in fahrer_name:
                        nachname, vorname = [teil.strip() for teil in fahrer_name.split(",", 1)]
                    else:
                        nachname, vorname = fahrer_name.strip(), ""

                    filename_part = sonder_dateien.get(
                        (nachname.lower(), vorname.lower()),
                        nachname.replace(" ", "_"),
                    )

                    alter_dateiname = f"KW{kw:02d}_{filename_part}.html".lower()
                    if "ch._holtz" in alter_dateiname or any(
                        wort in alter_dateiname for wort in ausschluss_stichwoerter
                    ):
                        continue

                    fahrer_wochen.add((kw, filename_part))
                    reihenfolge = 0

                    for tag_index in range(7):
                        tag_datum = start_sonntag + pd.Timedelta(days=tag_index)
                        tag_key = pd.Timestamp(tag_datum).date()
                        wochentag = wochentage_deutsch_map.get(
                            pd.Timestamp(tag_datum).strftime("%A"),
                            pd.Timestamp(tag_datum).strftime("%A"),
                        )

                        tag_eintraege = eintraege.get(tag_key, []) or ["– – –"]

                        for eintrag_text in tag_eintraege:
                            reihenfolge += 1
                            if " – " in eintrag_text:
                                uhrzeit_str, tour_str = eintrag_text.split(" – ", 1)
                            else:
                                uhrzeit_str, tour_str = "–", eintrag_text

                            uhrzeit_str = uhrzeit_str.strip() or "–"
                            tour_str = tour_str.strip() or "–"
                            if uhrzeit_str == "–":
                                uhrzeit_str = "–"
                            if tour_str == "–":
                                tour_str = "–"

                            datum_iso = pd.Timestamp(tag_datum).date().isoformat()
                            url = (
                                f"dienstplan_v4.html?kw={kw:02d}"
                                f"&fahrer={quote(filename_part, safe='')}"
                            )

                            row_key = (
                                kw, filename_part, fahrer_name,
                                datum_iso, uhrzeit_str, tour_str,
                            )
                            if row_key in bekannte_zeilen:
                                continue
                            bekannte_zeilen.add(row_key)

                            csv_rows.append({
                                "kw": kw,
                                "fahrer_key": filename_part,
                                "fahrer_name": fahrer_name,
                                "datum": datum_iso,
                                "wochentag": wochentag,
                                "uhrzeit": uhrzeit_str,
                                "tour": tour_str,
                                "reihenfolge": reihenfolge,
                                "url": url,
                            })

                dauer_datei = time.perf_counter() - datei_start
                status.info(
                    f"{uploaded_file.name}: {len(touren_liste):,} Tourenzeilen "
                    f"in {dauer_datei:.1f} Sekunden gelesen."
                    .replace(",", ".")
                )
                fortschritt.progress(datei_index / len(uploaded_files))

            if not csv_rows:
                st.warning("Es wurden keine Dienstplandaten erzeugt.")
                st.stop()

            status.info("Erzeuge gemeinsame HTML- und JavaScript-Datei …")
            with open(html_path, "w", encoding="utf-8") as handle:
                handle.write(generate_shared_html(css_styles))
            with open(js_path, "w", encoding="utf-8") as handle:
                handle.write(generate_shared_js())

            status.info("Bereinige die neu erzeugten Dienstplandaten …")
            new_csv_df = clean_and_sort_csv(pd.DataFrame(csv_rows))

            existing_csv_df = None
            retained_rows = 0
            replaced_rows = 0

            if all([FTP_HOST, FTP_USER, FTP_PASS]):
                status.info("Lade die vorhandene dienstplaene.csv …")
                merge_status, merge_message = download_existing_csv(existing_csv_path)

                if merge_status == "found":
                    status.info(merge_message)
                    existing_csv_df = read_dienstplan_csv(existing_csv_path)
                    st.info(
                        f"Vorhandene CSV geladen: {len(existing_csv_df):,} Zeilen."
                        .replace(",", ".")
                    )
                elif merge_status == "missing":
                    st.info(merge_message + " Es wird eine neue Gesamtdatei angelegt.")
                else:
                    st.error(merge_message)
                    st.error(
                        "Der Vorgang wurde abgebrochen, damit die bestehende CSV "
                        "nicht versehentlich ersetzt wird."
                    )
                    st.stop()
            else:
                st.warning(
                    "FTP-Zugangsdaten fehlen. Der Download enthält daher nur die "
                    "gerade hochgeladenen Wochen."
                )

            status.info("Füge neue und vorhandene Kalenderwochen zusammen …")
            csv_df, retained_rows, replaced_rows = merge_uploaded_weeks(
                existing_csv_df,
                new_csv_df,
                hochgeladene_wochen,
            )

            csv_df.to_csv(
                csv_path,
                sep=";",
                index=False,
                encoding="utf-8-sig",
                lineterminator="\n",
            )

            status.info("Erstelle ZIP-Datei …")
            with ZipFile(zip_path, "w") as zipf:
                zipf.write(html_path, arcname="dienstplan_v4.html")
                zipf.write(js_path, arcname="dienstplan_app_v4.js")
                zipf.write(csv_path, arcname="dienstplaene.csv")

            with open(zip_path, "rb") as handle:
                zip_bytes = handle.read()

            if auto_ftp:
                if not all([FTP_HOST, FTP_USER, FTP_PASS]):
                    st.warning("FTP-Zugangsdaten fehlen in der .env-Datei.")
                else:
                    status.info("Lade die drei Dateien auf den FTP-Server …")
                    upload_folder_to_ftp_with_progress(
                        tmpdir,
                        FTP_BASE_DIR,
                        allowed_names={
                            "dienstplan_v4.html",
                            "dienstplan_app_v4.js",
                            "dienstplaene.csv",
                        },
                    )

            gesamt_dauer = time.perf_counter() - gesamt_start
            status.success(f"Fertig nach {gesamt_dauer:.1f} Sekunden.")

            wochen_text = ", ".join(
                pd.Timestamp(woche).strftime("%d.%m.%Y")
                for woche in sorted(hochgeladene_wochen)
            )
            st.success(
                f"{len(uploaded_files)} Datei(en) verarbeitet. "
                f"{len(fahrer_wochen)} Fahrer-Woche(n) neu erstellt. "
                f"Die Gesamt-CSV enthält jetzt {len(csv_df):,} Zeilen."
                .replace(",", ".")
            )

            if existing_csv_df is not None:
                st.info(
                    f"Ersetzte Planwoche(n), jeweils ab Sonntag: {wochen_text}. "
                    f"{replaced_rows:,} alte Zeilen wurden ersetzt und "
                    f"{retained_rows:,} Zeilen aus anderen Wochen beibehalten."
                    .replace(",", ".")
                )

            st.download_button(
                "ZIP mit HTML-, JavaScript- und CSV-Datei herunterladen",
                data=zip_bytes,
                file_name="gesamt_export.zip",
                mime="application/zip",
            )

    except Exception as exc:
        st.exception(exc)
